import time

start_time = time.time()

import csv
import openpyxl
from decimal import Decimal

report_month = 202204


# ACCESSING EXCEL FILE
# ACCESSING EXCEL FILE
# ACCESSING EXCEL FILE

wb = openpyxl.load_workbook(filename=r'C:\Users\jazz.ooi\Desktop\proj\Sharpinvest\report_20220819.xlsx')
# filepath of xls data; convert to xlsx for maximum compatability


#corr_data = wb['Sheet1']
corr_data = wb[wb.sheetnames[0]]


xl_rows = corr_data.max_row

xl_filenames = []
xl_202207 = []
xl_202206 = []
xl_202205 = []
xl_202204 = []

for i in range(2, xl_rows + 1):
    company_id = corr_data.cell(row=i, column=1).value
    prod_id = corr_data.cell(row=i, column=2).value
    xl_filenames.append(company_id + "_" + prod_id + ".pdf")
    company_202207 = corr_data.cell(row=i, column=5).value
    company_202206 = corr_data.cell(row=i, column=6).value
    company_202205 = corr_data.cell(row=i, column=7).value
    company_202204 = corr_data.cell(row=i, column=8).value
    xl_202207.append(Decimal(company_202207))
    xl_202206.append(Decimal(company_202206))
    xl_202205.append(Decimal(company_202205))
    xl_202204.append(Decimal(company_202204))

xl_dict = {}

for i in range(0, len(xl_filenames)):
    xl_dict[xl_filenames[i]] = {202204: xl_202204[i], 202205: xl_202205[i], 202206: xl_202206[i], 202207: xl_202207[i]}


# ACCESSING CSV FILE
# ACCESSING CSV FILE
# ACCESSING CSV FILE

csv_dir_202204 = "C://Users//jazz.ooi//Desktop//proj//Sharpinvest//csv_outputs//output_202204.csv"
csv_dir_202205 = "C://Users//jazz.ooi//Desktop//proj//Sharpinvest//csv_outputs//output_202205.csv"
csv_dir_202207 = "C://Users//jazz.ooi//Desktop//proj//Sharpinvest//csv_outputs//output_202207.csv"

with open(csv_dir_202205, 'r', encoding='utf-16') as csvfile202205:
    test_data_202205 = csv.DictReader(csvfile202205, delimiter=',', dialect="excel")
    test_data_202205 = list(test_data_202205)

with open(csv_dir_202204, 'r', encoding='utf-16') as csvfile202204:
    test_data_202204 = csv.DictReader(csvfile202204, delimiter=',', dialect="excel")
    test_data_202204 = list(test_data_202204)

with open(csv_dir_202207, 'r', encoding='utf-16') as csvfile202207:
    test_data_202207 = csv.DictReader(csvfile202207, delimiter=',', dialect="excel")
    test_data_202207 = list(test_data_202207)

delta_202205 = {}
for i in range(len(test_data_202205)):
    filename = test_data_202205[i]['filename']
    aum_csv = test_data_202205[i]['AUM']
    if filename in xl_dict:
        aum_xl = xl_dict[filename][202205] #or 202206?
        if aum_csv != "AUM not found" and aum_xl != 0:
            delta_202205[filename] = abs(Decimal(aum_csv) - Decimal(aum_xl)) #abs
        elif aum_csv == "AUM not found":
            print(filename, "AUM not found")
#note: with aum_xl != 0, length is 358; without its 360
print('')
total_202205 = 0
for key in delta_202205:
    if delta_202205[key] != 0:
        print('202205:', key, delta_202205[key])
    total_202205 += delta_202205[key]
print(total_202205, "\n\n")

print("=================202204========================")
delta_202204 = {}
for i in range(len(test_data_202204)):
    filename = test_data_202204[i]['filename']
    aum_csv = test_data_202204[i]['AUM']
    if filename in xl_dict:
        aum_xl = xl_dict[filename][202204] #or 202205?
        if aum_csv != "AUM not found": #and aum_xl != 0:
            delta_202204[filename] = abs(Decimal(aum_csv) - Decimal(aum_xl)) #abs
        elif aum_csv == "AUM not found":
            print(filename, "AUM not found")
#note: witwithhout aum_xl != 0, length is 363; without its 365
print('')
total_202204 = 0
for key in delta_202204:
    if delta_202204[key] != 0:
        print('202204:', key, delta_202204[key])
    total_202204 += delta_202204[key]
print(total_202204)


print("=================202207========================")

delta_202207 = {}
for i in range(len(test_data_202207)):
    filename = test_data_202207[i]['filename']
    aum_csv = test_data_202207[i]['AUM']
    if filename in xl_dict:
        aum_xl = xl_dict[filename][202207] #or 202206?
        if aum_csv != "AUM not found" and aum_xl != 0:
            delta_202207[filename] = abs(int(Decimal(aum_csv)) - int(Decimal(aum_xl))) #abs
        elif aum_csv == "AUM not found":
            print("202207:", filename, "AUM not found")
#note: with aum_xl != 0, length is 358; without its 360
print('')
total_202207 = 0
for key in delta_202207:
    if delta_202207[key] != 0:
        print('202207:', key, delta_202207[key])
    total_202207 += delta_202207[key]
print(total_202207, "\n\n")

print("\n\n time elapsed: {:.2f}s".format(time.time() - start_time))
